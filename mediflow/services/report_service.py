"""Reporting service: read-only aggregate reports across the clinic's data.

Each report is returned as a :class:`ReportResult` (summary metrics + an optional
detail table) which the UI renders and can export to Excel.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from sqlalchemy import func, select

from mediflow.core.constants import (
    AppointmentStatus,
    Gender,
    InvoiceStatus,
    LabRequestStatus,
)
from mediflow.data.base import local_day_bounds_utc, local_month_start_utc, local_today, utcnow
from mediflow.data.database import Database
from mediflow.data.models.appointment import Appointment
from mediflow.data.models.billing import Invoice
from mediflow.data.models.inventory import InventoryItem
from mediflow.data.models.laboratory import LabRequest
from mediflow.data.models.patient import Patient
from mediflow.data.models.pharmacy import Medication


@dataclass(slots=True)
class ReportResult:
    key: str
    title: str
    summary: list[tuple[str, object]] = field(default_factory=list)
    columns: list[str] = field(default_factory=list)
    rows: list[list[str]] = field(default_factory=list)


# (key, translatable title) — the catalogue shown in the Reports view.
REPORTS = [
    ("patients", "Patient summary"),
    ("appointments", "Appointment summary"),
    ("revenue", "Revenue summary"),
    ("pharmacy", "Pharmacy stock"),
    ("inventory", "Inventory stock"),
    ("lab", "Laboratory summary"),
]


class ReportService:
    def __init__(self, db: Database):
        self._db = db

    def generate(self, key: str) -> ReportResult:
        return {
            "patients": self._patients,
            "appointments": self._appointments,
            "revenue": self._revenue,
            "pharmacy": self._pharmacy,
            "inventory": self._inventory,
            "lab": self._lab,
        }[key]()

    # -- individual reports -------------------------------------------------
    def _patients(self) -> ReportResult:
        # created_at is UTC-stamped; bucket by the local month/day.
        month_start = local_month_start_utc()
        day_start, day_end = local_day_bounds_utc()
        with self._db.unit_of_work() as session:
            def count(*where) -> int:
                stmt = select(func.count()).select_from(Patient).where(
                    Patient.is_deleted.is_(False), *where)
                return int(session.execute(stmt).scalar_one())

            summary = [
                ("Total patients", count()),
                ("Male", count(Patient.gender == Gender.MALE.value)),
                ("Female", count(Patient.gender == Gender.FEMALE.value)),
                ("Registered this month", count(Patient.created_at >= month_start)),
                ("Registered today", count(Patient.created_at >= day_start,
                                           Patient.created_at < day_end)),
            ]
        return ReportResult("patients", "Patient summary", summary=summary)

    def _appointments(self) -> ReportResult:
        # scheduled_start is local wall-clock; the local month needs no shift.
        today = local_today()
        month_start = datetime(today.year, today.month, 1)  # noqa: DTZ001  (wall-clock domain — see mediflow/data/base.py)
        with self._db.unit_of_work() as session:
            def count(*where) -> int:
                stmt = select(func.count()).select_from(Appointment).where(
                    Appointment.is_deleted.is_(False),
                    Appointment.scheduled_start >= month_start, *where)
                return int(session.execute(stmt).scalar_one())

            summary = [
                ("Total", count()),
                ("Booked", count(Appointment.status == AppointmentStatus.BOOKED.value)),
                ("Completed", count(Appointment.status == AppointmentStatus.COMPLETED.value)),
                ("Cancelled", count(Appointment.status == AppointmentStatus.CANCELLED.value)),
                ("No show", count(Appointment.status == AppointmentStatus.NO_SHOW.value)),
            ]
        return ReportResult("appointments", "Appointment summary", summary=summary)

    def _revenue(self) -> ReportResult:
        with self._db.unit_of_work() as session:
            stmt = (
                select(Invoice)
                .where(Invoice.is_deleted.is_(False),
                       Invoice.status != InvoiceStatus.CANCELLED.value)
                .order_by(Invoice.issued_at.desc())
            )
            invoices = session.execute(stmt).scalars().all()
            invoiced = round(sum(float(i.total or 0) for i in invoices), 2)
            collected = round(sum(float(i.paid_amount or 0) for i in invoices), 2)
            summary = [
                ("Invoices", len(invoices)),
                ("Invoiced", invoiced),
                ("Collected", collected),
                ("Outstanding", round(invoiced - collected, 2)),
            ]
            rows = [
                [f"‎{i.invoice_number}‎",
                 i.patient.full_name if i.patient else "—",
                 f"{float(i.total or 0):.2f}",
                 f"{float(i.paid_amount or 0):.2f}"]
                for i in invoices[:50]
            ]
        return ReportResult("revenue", "Revenue summary", summary=summary,
                            columns=["Invoice", "Patient", "Total", "Paid"], rows=rows)

    def _pharmacy(self) -> ReportResult:
        with self._db.unit_of_work() as session:
            meds = session.execute(
                select(Medication).where(Medication.is_deleted.is_(False))
            ).scalars().all()
            low_rows, low, out, expiring = [], 0, 0, 0
            today = local_today()  # expiry dates are local calendar dates
            for m in meds:
                live = [b for b in m.batches if not b.is_deleted and b.quantity > 0]
                stock = sum(b.quantity for b in live)
                expiries = [b.expiry_date for b in live if b.expiry_date]
                is_expiring = bool(expiries) and (min(expiries) - today).days <= 30
                if stock <= 0:
                    out += 1
                if stock <= m.reorder_level:
                    low += 1
                if is_expiring:
                    expiring += 1
                if stock <= m.reorder_level or is_expiring:
                    low_rows.append([m.name, str(stock), str(m.reorder_level)])
            summary = [
                ("Medications", len(meds)),
                ("Low stock", low),
                ("Out of stock", out),
                ("Expiring", expiring),
            ]
        return ReportResult("pharmacy", "Pharmacy stock", summary=summary,
                            columns=["Medication", "Stock", "Reorder"], rows=low_rows)

    def _inventory(self) -> ReportResult:
        with self._db.unit_of_work() as session:
            items = session.execute(
                select(InventoryItem).where(InventoryItem.is_deleted.is_(False))
            ).scalars().all()
            low_rows, low, out = [], 0, 0
            for i in items:
                if i.quantity <= 0:
                    out += 1
                if i.quantity <= i.reorder_level:
                    low += 1
                    low_rows.append([i.name, str(i.quantity), str(i.reorder_level)])
            summary = [
                ("Items", len(items)),
                ("Low stock", low),
                ("Out of stock", out),
            ]
        return ReportResult("inventory", "Inventory stock", summary=summary,
                            columns=["Item", "Quantity", "Reorder"], rows=low_rows)

    def _lab(self) -> ReportResult:
        with self._db.unit_of_work() as session:
            def count(*where) -> int:
                stmt = select(func.count()).select_from(LabRequest).where(
                    LabRequest.is_deleted.is_(False), *where)
                return int(session.execute(stmt).scalar_one())

            summary = [
                ("Total", count()),
                ("Requested", count(LabRequest.status == LabRequestStatus.REQUESTED.value)),
                ("In progress", count(LabRequest.status == LabRequestStatus.IN_PROGRESS.value)),
                ("Completed", count(LabRequest.status == LabRequestStatus.COMPLETED.value)),
                ("Cancelled", count(LabRequest.status == LabRequestStatus.CANCELLED.value)),
            ]
        return ReportResult("lab", "Laboratory summary", summary=summary)


def export_excel(result: ReportResult, directory: Path,
                 *, title: str, summary_labels: dict[str, str],
                 column_labels: dict[str, str]) -> Path:
    """Write the report to an .xlsx file; return the created path.

    ``title``/``summary_labels``/``column_labels`` are the already-translated
    strings supplied by the UI so the export matches the on-screen language.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = result.key[:31]

    bold = Font(bold=True)
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    for label_key, value in result.summary:
        ws.append([summary_labels.get(label_key, label_key), value])
    if result.columns:
        ws.append([])
        header = [column_labels.get(c, c) for c in result.columns]
        ws.append(header)
        for cell in ws[ws.max_row]:
            cell.font = bold
        for row in result.rows:
            ws.append(row)

    for column_cells in ws.columns:
        width = max((len(str(c.value)) for c in column_cells if c.value is not None), default=10)
        ws.column_dimensions[column_cells[0].column_letter].width = min(width + 4, 50)

    directory.mkdir(parents=True, exist_ok=True)
    stamp = utcnow().strftime("%Y%m%d_%H%M%S")
    path = directory / f"{result.key}_{stamp}.xlsx"
    wb.save(path)
    return path
